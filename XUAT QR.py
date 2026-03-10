import os
import re
import zipfile
import shutil
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import load_workbook

# ====== CẤU HÌNH ======
EXCEL_FILE = "MAU QUET QR.xlsx"
SHEET_NAME = "DanhSach"
OUTPUT_FOLDER = "output_qr"

# Cột trong Excel
COL_NAME = 2   # cột B = Họ tên
COL_CCCD = 4   # cột D = Số căn cước


def safe_filename(text):
    text = str(text).strip()
    text = re.sub(r'[\\/:*?"<>|]', '_', text)
    text = re.sub(r'\s+', '_', text)
    return text


def normalize_name(name):
    name = str(name).strip().upper()
    name = re.sub(r'\s+', '_', name)
    name = re.sub(r'[\\/:*?"<>|]', '_', name)
    return name


def get_sheet_and_drawing_map(xlsx_path):
    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_rel_doc = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns_rel_pkg = "http://schemas.openxmlformats.org/package/2006/relationships"

    result = {}

    with zipfile.ZipFile(xlsx_path, "r") as z:
        workbook_xml = ET.fromstring(z.read("xl/workbook.xml"))
        workbook_rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))

        rel_map = {}
        for rel in workbook_rels.findall("{%s}Relationship" % ns_rel_pkg):
            rid = rel.attrib["Id"]
            target = rel.attrib["Target"]
            if not target.startswith("xl/"):
                target = "xl/" + target
            rel_map[rid] = target

        sheets = workbook_xml.find("{%s}sheets" % ns_main)
        for sh in sheets.findall("{%s}sheet" % ns_main):
            sheet_name = sh.attrib["name"]
            rid = sh.attrib["{%s}id" % ns_rel_doc]
            sheet_path = rel_map[rid]

            drawing_path = None
            rels_path = os.path.dirname(sheet_path).replace("\\", "/") + "/_rels/" + os.path.basename(sheet_path) + ".rels"

            if rels_path in z.namelist():
                sheet_xml = ET.fromstring(z.read(sheet_path))
                drawing_el = sheet_xml.find("{%s}drawing" % ns_main)

                if drawing_el is not None:
                    drawing_rid = drawing_el.attrib["{%s}id" % ns_rel_doc]
                    rels_xml = ET.fromstring(z.read(rels_path))

                    for rel in rels_xml.findall("{%s}Relationship" % ns_rel_pkg):
                        if rel.attrib["Id"] == drawing_rid:
                            target = rel.attrib["Target"]
                            base_dir = os.path.dirname(sheet_path).replace("\\", "/")
                            drawing_path = os.path.normpath(os.path.join(base_dir, target)).replace("\\", "/")
                            break

            result[sheet_name] = {
                "sheet_path": sheet_path,
                "drawing_path": drawing_path
            }

    return result


def get_image_relations(zip_file, drawing_xml_path):
    ns_rel_pkg = "http://schemas.openxmlformats.org/package/2006/relationships"

    drawing_dir = os.path.dirname(drawing_xml_path).replace("\\", "/")
    drawing_name = os.path.basename(drawing_xml_path)
    rels_path = drawing_dir + "/_rels/" + drawing_name + ".rels"

    rel_map = {}
    if rels_path not in zip_file.namelist():
        return rel_map

    rels_xml = ET.fromstring(zip_file.read(rels_path))
    for rel in rels_xml.findall("{%s}Relationship" % ns_rel_pkg):
        rid = rel.attrib["Id"]
        target = rel.attrib["Target"]
        image_path = os.path.normpath(os.path.join(drawing_dir, target)).replace("\\", "/")
        rel_map[rid] = image_path

    return rel_map


def extract_images_by_row(xlsx_path, sheet_name, temp_folder):
    ns = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }

    mapping = get_sheet_and_drawing_map(xlsx_path)
    drawing_xml_path = mapping[sheet_name]["drawing_path"]

    if not drawing_xml_path:
        raise Exception("Sheet không có ảnh.")

    results = []

    with zipfile.ZipFile(xlsx_path, "r") as z:
        drawing_xml = ET.fromstring(z.read(drawing_xml_path))
        rel_map = get_image_relations(z, drawing_xml_path)

        anchors = drawing_xml.findall("xdr:oneCellAnchor", ns) + drawing_xml.findall("xdr:twoCellAnchor", ns)

        if not os.path.exists(temp_folder):
            os.makedirs(temp_folder)

        stt = 1
        for anchor in anchors:
            from_el = anchor.find("xdr:from", ns)
            if from_el is None:
                continue

            row_el = from_el.find("xdr:row", ns)
            if row_el is None:
                continue

            row_excel = int(row_el.text) + 1

            pic = anchor.find("xdr:pic", ns)
            if pic is None:
                continue

            blip = pic.find(".//a:blip", ns)
            if blip is None:
                continue

            embed = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
            if not embed or embed not in rel_map:
                continue

            image_in_zip = rel_map[embed]
            if image_in_zip not in z.namelist():
                continue

            ext = Path(image_in_zip).suffix.lower()
            if ext == "":
                ext = ".png"

            temp_path = os.path.join(temp_folder, "temp_%s%s" % (stt, ext))

            with z.open(image_in_zip) as src, open(temp_path, "wb") as dst:
                shutil.copyfileobj(src, dst)

            results.append({
                "row": row_excel,
                "temp_path": temp_path
            })
            stt += 1

    return results


def main():
    if not os.path.exists(EXCEL_FILE):
        print("Không tìm thấy file Excel:", EXCEL_FILE)
        return

    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)

    temp_folder = os.path.join(OUTPUT_FOLDER, "__temp__")

    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    images = extract_images_by_row(EXCEL_FILE, SHEET_NAME, temp_folder)

    ok = 0
    skip = 0

    for item in images:
        row = item["row"]

        # Bỏ qua nếu ảnh ở dòng tiêu đề
        if row <= 1:
            try:
                os.remove(item["temp_path"])
            except:
                pass
            skip += 1
            continue

        name_value = ws.cell(row=row, column=COL_NAME).value
        cccd_value = ws.cell(row=row, column=COL_CCCD).value

        if not name_value or not cccd_value:
            try:
                os.remove(item["temp_path"])
            except:
                pass
            skip += 1
            continue

        name_text = normalize_name(name_value)
        cccd_text = str(cccd_value).strip()

        filename = safe_filename(name_text + "_" + cccd_text + ".png")
        final_path = os.path.join(OUTPUT_FOLDER, filename)

        if os.path.exists(final_path):
            base = Path(filename).stem
            ext = Path(filename).suffix
            n = 2
            while True:
                test_name = base + "_" + str(n) + ext
                test_path = os.path.join(OUTPUT_FOLDER, test_name)
                if not os.path.exists(test_path):
                    final_path = test_path
                    break
                n += 1

        shutil.move(item["temp_path"], final_path)
        ok += 1

    if os.path.exists(temp_folder):
        try:
            os.rmdir(temp_folder)
        except:
            pass

    print("====================================")
    print("Xuất ảnh thành công:", ok)
    print("Bỏ qua:", skip)
    print("Thư mục output:", os.path.abspath(OUTPUT_FOLDER))
    print("====================================")


if __name__ == "__main__":
    main()
